"""
Session Accumulator - Acumulador de comprobantes para la sesión actual.
Mantiene los datos en memoria y permite exportación manual a Excel.
"""
import os
import json
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.paths import get_app_data_dir

logger = logging.getLogger(__name__)


class SessionAccumulator:
    """
    Acumulador de comprobantes para la sesión actual.
    Mantiene los datos en memoria con persistencia a disco.
    """
    
    def __init__(self, persistence_file: Optional[str] = None):
        """
        Inicializa el acumulador.
        
        Args:
            persistence_file: Ruta al archivo JSON para persistencia (opcional)
        """
        if persistence_file is None:
            persistence_file = os.path.join(get_app_data_dir(), "session_data.json")
        
        self.persistence_file = persistence_file
        self.entries: List[Dict[str, Any]] = []
        self.session_start: str = datetime.now().isoformat()
        self.export_history: List[Dict[str, Any]] = []
        
        # Cargar datos persistidos si existen
        self._load_from_disk()
    
    def _load_from_disk(self) -> None:
        """Carga los datos persistidos desde disco."""
        try:
            if os.path.exists(self.persistence_file):
                with open(self.persistence_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.entries = data.get('entries', [])
                    self.session_start = data.get('session_start', self.session_start)
                    self.export_history = data.get('export_history', [])
                    logger.info(f"Cargados {len(self.entries)} comprobantes de la sesión anterior")
        except Exception as e:
            logger.error(f"Error cargando datos de sesión: {e}")
            self.entries = []
    
    def _save_to_disk(self) -> None:
        """Persiste los datos a disco."""
        try:
            os.makedirs(os.path.dirname(self.persistence_file), exist_ok=True)
            data = {
                'entries': self.entries,
                'session_start': self.session_start,
                'export_history': self.export_history,
                'last_updated': datetime.now().isoformat()
            }
            with open(self.persistence_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logger.error(f"Error guardando datos de sesión: {e}")
    
    def add_entry(self, data: Dict[str, Any]) -> None:
        """
        Agrega un comprobante al acumulador.
        
        Args:
            data: Diccionario con los datos del comprobante
        """
        # Recargar desde disco antes de agregar para detectar resets
        self._load_from_disk()
        
        entry = {
            'timestamp': datetime.now().isoformat(),
            'archivo': data.get('archivo', 'Sin nombre'),
            'fuente': data.get('fuente', 'WhatsApp'),
            'fecha_operacion': data.get('fecha_operacion', ''),
            'monto': self._parse_monto(data.get('monto', 0)),
            'banco_origen': data.get('banco_origen', ''),
            'banco_destino': data.get('banco_destino', ''),
            'cbu_origen': data.get('cbu_origen', ''),
            'cbu_destino': data.get('cbu_destino', ''),
            'ordenante': data.get('ordenante', ''),
            'receptor_nombre': data.get('receptor_nombre', ''),
            'receptor_cuit': data.get('receptor_cuit', ''),
            'numero_comprobante': data.get('numero_comprobante', ''),
            'estado': 'Procesado',
            'whatsapp_from': data.get('whatsapp_from', ''),
            'cuenta_destino': data.get('cuenta_destino', 'Sin especificar')
        }
        self.entries.append(entry)
        self._save_to_disk()
        logger.info(f"Comprobante agregado: {entry['archivo']} - ${entry['monto']}")
    
    def _parse_monto(self, monto) -> float:
        """Convierte monto a float."""
        if isinstance(monto, (int, float)):
            return float(monto)
        if isinstance(monto, str):
            # Limpiar el string de monto
            cleaned = monto.replace('$', '').replace('.', '').replace(',', '.').strip()
            try:
                return float(cleaned)
            except ValueError:
                return 0.0
        return 0.0
    
    def reload(self) -> None:
        """Recarga los datos desde disco (para sincronizar entre procesos)."""
        self._load_from_disk()
    
    def get_count(self) -> int:
        """Retorna la cantidad de comprobantes en la sesión."""
        return len(self.entries)
    
    def get_total_amount(self) -> float:
        """Retorna la suma total de montos."""
        return sum(entry.get('monto', 0) for entry in self.entries)
    
    def get_recent_entries(self, count: int = 10) -> List[Dict[str, Any]]:
        """
        Retorna los últimos N comprobantes.
        
        Args:
            count: Cantidad de comprobantes a retornar
            
        Returns:
            Lista de los últimos comprobantes (más recientes primero)
        """
        return list(reversed(self.entries[-count:]))
    
    def get_session_info(self) -> Dict[str, Any]:
        """Retorna información de la sesión actual."""
        return {
            'session_start': self.session_start,
            'count': self.get_count(),
            'total_amount': self.get_total_amount(),
            'last_entry': self.entries[-1] if self.entries else None,
            'export_count': len(self.export_history)
        }
    
    def export_to_excel(self, output_dir: Optional[str] = None, 
                        filename_prefix: str = "comprobantes") -> Dict[str, Any]:
        """
        Exporta los datos acumulados a un archivo Excel.
        
        Args:
            output_dir: Directorio de salida (default: AppData)
            filename_prefix: Prefijo del nombre de archivo
            
        Returns:
            Dict con resultado de la exportación
        """
        if not self.entries:
            return {
                'success': False,
                'error': 'No hay comprobantes para exportar'
            }
        
        try:
            if output_dir is None:
                output_dir = get_app_data_dir()
            
            os.makedirs(output_dir, exist_ok=True)
            
            # Generar nombre con timestamp
            timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            filename = f"{filename_prefix}_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Comprobantes"
            
            # Headers
            headers = [
                "Fecha/Hora", "Archivo", "Fuente", "Fecha Operación",
                "Monto", "Banco Origen", "Banco Destino", "CBU Origen",
                "CBU Destino", "Ordenante", "Receptor", "CUIT Receptor",
                "Nº Comprobante", "Estado", "WhatsApp", "Cuenta Destino"
            ]
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0D5C5A", end_color="0D5C5A", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Escribir headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            
            # Escribir datos
            for row_idx, entry in enumerate(self.entries, 2):
                row_data = [
                    entry.get('timestamp', ''),
                    entry.get('archivo', ''),
                    entry.get('fuente', ''),
                    entry.get('fecha_operacion', ''),
                    entry.get('monto', 0),
                    entry.get('banco_origen', ''),
                    entry.get('banco_destino', ''),
                    entry.get('cbu_origen', ''),
                    entry.get('cbu_destino', ''),
                    entry.get('ordenante', ''),
                    entry.get('receptor_nombre', ''),
                    entry.get('receptor_cuit', ''),
                    entry.get('numero_comprobante', ''),
                    entry.get('estado', ''),
                    entry.get('whatsapp_from', ''),
                    entry.get('cuenta_destino', '')
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border
                    # Formatear monto como moneda
                    if col == 5:  # Columna Monto
                        cell.number_format = '"$"#,##0.00'
            
            # Ajustar anchos de columna
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Fila de totales
            total_row = len(self.entries) + 2
            ws.cell(row=total_row, column=4, value="TOTAL:").font = Font(bold=True)
            total_cell = ws.cell(row=total_row, column=5, value=self.get_total_amount())
            total_cell.font = Font(bold=True)
            total_cell.number_format = '"$"#,##0.00'
            
            wb.save(filepath)
            
            # Registrar en historial
            export_record = {
                'timestamp': datetime.now().isoformat(),
                'filepath': filepath,
                'count': len(self.entries),
                'total_amount': self.get_total_amount()
            }
            self.export_history.append(export_record)
            self._save_to_disk()
            
            logger.info(f"Exportados {len(self.entries)} comprobantes a {filepath}")
            
            return {
                'success': True,
                'filepath': filepath,
                'count': len(self.entries),
                'total_amount': self.get_total_amount()
            }
            
        except Exception as e:
            logger.error(f"Error exportando a Excel: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def reset(self, export_first: bool = True, output_dir: Optional[str] = None) -> Dict[str, Any]:
        """
        Reinicia el acumulador, opcionalmente exportando primero.
        
        Args:
            export_first: Si es True, exporta antes de limpiar
            output_dir: Directorio para la exportación
            
        Returns:
            Dict con resultado de la operación
        """
        export_result = None
        
        if export_first and self.entries:
            export_result = self.export_to_excel(output_dir)
            if not export_result.get('success'):
                return {
                    'success': False,
                    'error': f"Error al exportar: {export_result.get('error')}"
                }
        
        # Limpiar datos
        previous_count = len(self.entries)
        previous_total = self.get_total_amount()
        
        self.entries = []
        self.session_start = datetime.now().isoformat()
        self._save_to_disk()
        
        logger.info(f"Sesión reiniciada. Se procesaron {previous_count} comprobantes por ${previous_total}")
        
        return {
            'success': True,
            'previous_count': previous_count,
            'previous_total': previous_total,
            'export_result': export_result
        }
    
    def get_export_history(self, count: int = 5) -> List[Dict[str, Any]]:
        """Retorna el historial de las últimas N exportaciones."""
        return list(reversed(self.export_history[-count:]))


# Instancia global para uso compartido
_accumulator_instance: Optional[SessionAccumulator] = None


def get_accumulator() -> SessionAccumulator:
    """Obtiene la instancia global del acumulador."""
    global _accumulator_instance
    if _accumulator_instance is None:
        _accumulator_instance = SessionAccumulator()
    return _accumulator_instance


def reset_accumulator() -> None:
    """Reinicia la instancia global del acumulador."""
    global _accumulator_instance
    _accumulator_instance = None
