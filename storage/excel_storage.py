"""
Almacenamiento de transferencias en Excel local usando openpyxl.
"""
import os
import threading
import time
import logging
from datetime import datetime
from typing import Optional, List, Dict, Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from app.paths import resolve_appdata_path

# Logger para este módulo
logger = logging.getLogger(__name__)

# Cola de entradas pendientes (cuando el Excel está bloqueado)
_pending_entries: List[Dict[str, Any]] = []
_pending_lock = threading.Lock()
_retry_thread: Optional[threading.Thread] = None
_retry_running = False


# Headers del Excel - Actualizado con columnas de Receptor y WhatsApp
HEADERS = [
    "Archivo",              # Nombre del archivo procesado
    "Fuente",               # De dónde viene (WhatsApp, etc.)
    "Fecha Operación",      # Fecha del comprobante
    "Monto",                # Monto de la transferencia
    "Emisor Nombre",        # Nombre del que envía
    "CUIT Emisor",          # CUIT/CUIL del emisor
    "CBU/CVU Emisor",       # CBU o CVU del emisor
    "Banco Emisor",         # Banco del emisor
    "Receptor Nombre",      # Nombre del que recibe
    "CUIT Receptor",        # CUIT/CUIL del receptor
    "CBU/CVU Receptor",     # CBU o CVU del receptor
    "Banco Receptor",       # Banco del receptor
    "Referencia",           # Número de referencia
    "Concepto",             # Concepto/descripción
    "Confianza",            # Nivel de confianza de extracción
    "Errores",              # Errores de validación
    "WhatsApp"              # Link clickeable a WhatsApp Web
]

# Estilos
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
DUPLICATE_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")


def _crear_excel_con_headers(ruta: str) -> Workbook:
    """Crea un nuevo archivo Excel con los headers formateados."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transferencias"
    
    # Agregar headers
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    
    # Ajustar anchos de columna (17 columnas)
    column_widths = [25, 12, 18, 15, 25, 15, 25, 18, 25, 15, 25, 18, 20, 25, 12, 30, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    wb.save(ruta)
    return wb


def _detectar_duplicado(ws, fecha_deposito: str, monto: float) -> bool:
    """Detecta si ya existe una transferencia con la misma fecha y monto."""
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        if row[1] and row[2]:  # Fecha Depósito y Monto
            row_fecha = str(row[1]).strip()
            try:
                row_monto = float(str(row[2]).replace('$', '').replace(',', '').strip())
                if row_fecha == fecha_deposito and abs(row_monto - monto) < 1.0:
                    return True
            except (ValueError, TypeError):
                continue
    return False


def guardar_en_excel(
    datos: dict,
    ruta_excel: str,
    whatsapp_from: str = "",
    timestamp_recepcion: str = "",
    cuenta_destino: str = "Cuenta Desconocida"
) -> dict:
    """
    Guarda una transferencia en un archivo Excel local.
    
    Args:
        datos: Diccionario con los datos extraídos del comprobante
        ruta_excel: Ruta al archivo Excel
        whatsapp_from: Número de WhatsApp del remitente
        timestamp_recepcion: Timestamp de recepción del comprobante
        cuenta_destino: Nombre de la cuenta destino identificada
        
    Returns:
        Dict con resultado de la operación
    """
    try:
        # Resolver ruta (relativa a AppData)
        ruta_excel = resolve_appdata_path(ruta_excel, fallback_name="transferencias.xlsx")
        
        # Asegurar que el archivo tenga extensión .xlsx
        if not ruta_excel.lower().endswith('.xlsx'):
            ruta_excel += '.xlsx'

        # Crear directorio si no existe
        directorio = os.path.dirname(ruta_excel)
        if directorio and not os.path.exists(directorio):
            os.makedirs(directorio)
        
        # Cargar o crear archivo
        if os.path.exists(ruta_excel):
            wb = load_workbook(ruta_excel)
            ws = wb.active
        else:
            wb = _crear_excel_con_headers(ruta_excel)
            ws = wb.active
        
        # Formatear fecha recepción
        try:
            fecha_recepcion = datetime.fromisoformat(timestamp_recepcion.replace("Z", "+00:00"))
            fecha_formateada = fecha_recepcion.strftime("%d/%m/%Y %H:%M")
        except:
            fecha_formateada = timestamp_recepcion or datetime.now().strftime("%d/%m/%Y %H:%M")
        
        # Preparar datos
        monto = datos.get("monto_numerico", 0)
        fecha_deposito = str(datos.get("fecha_operacion", "")).strip()
        
        # Lógica para nombre de emisor (Manejo de Depósitos)
        emisor_nombre = datos.get("emisor_nombre", "")
        if not emisor_nombre:
            concepto = datos.get("concepto", "").lower()
            if "deposito" in concepto or "efectivo" in concepto:
                emisor_nombre = "DEPÓSITO EN EFECTIVO"
        
        # Convertir confianza a etiqueta
        confianza_val = datos.get("confianza", 0)
        confianza_str = "OPTIMA" if confianza_val >= 0.90 else "REVEER"
        
        # Detectar duplicado
        es_duplicado = _detectar_duplicado(ws, fecha_deposito, monto)
        
        # Preparar número de WhatsApp
        numero_wa = whatsapp_from.replace("@c.us", "") if whatsapp_from else ""
        
        # Crear fila con las 16 columnas
        # Errores de validación
        errores_list = datos.get("errores", [])
        errores_str = "; ".join(errores_list) if isinstance(errores_list, list) else str(errores_list or "")
        
        fila = [
            datos.get("archivo_origen", ""),      # A: Archivo
            "WhatsApp" if numero_wa else "API",   # B: Fuente
            fecha_deposito,                       # C: Fecha Operación
            monto,                                # D: Monto
            emisor_nombre,                        # E: Emisor Nombre
            datos.get("emisor_cuil", ""),         # F: CUIT Emisor
            datos.get("emisor_cbu", ""),          # G: CBU/CVU Emisor
            datos.get("banco_emisor", ""),        # H: Banco Emisor
            datos.get("receptor_nombre", ""),     # I: Receptor Nombre
            datos.get("receptor_cuil", ""),       # J: CUIT Receptor
            datos.get("receptor_cbu", ""),        # K: CBU/CVU Receptor
            datos.get("banco_receptor", ""),      # L: Banco Receptor
            datos.get("referencia", ""),          # M: Referencia
            datos.get("concepto", ""),            # N: Concepto
            confianza_str,                        # O: Confianza
            errores_str,                          # P: Errores
            numero_wa                             # Q: WhatsApp (se agrega hipervínculo después)
        ]
        
        # Agregar fila
        ws.append(fila)
        nueva_fila = ws.max_row
        
        # Marcar como duplicado si corresponde
        if es_duplicado:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=nueva_fila, column=col).fill = DUPLICATE_FILL
        
        # Agregar hipervínculo de WhatsApp (columna Q = 17)
        if numero_wa:
            wa_cell = ws.cell(row=nueva_fila, column=17)
            # Crear link a WhatsApp Web
            wa_link = f"https://wa.me/{numero_wa}"
            wa_cell.hyperlink = wa_link
            wa_cell.style = "Hyperlink"
        
        # Guardar
        wb.save(ruta_excel)
        
        return {
            "success": True,
            "message": "Guardado en Excel" + (" (Duplicado)" if es_duplicado else ""),
            "ruta": ruta_excel,
            "fila": nueva_fila,
            "es_duplicado": es_duplicado
        }
        
    except PermissionError as e:
        # El archivo está bloqueado (probablemente abierto en Excel)
        # Agregar a la cola de reintentos
        entry = {
            "datos": datos,
            "ruta_excel": ruta_excel,
            "whatsapp_from": whatsapp_from,
            "timestamp_recepcion": timestamp_recepcion,
            "cuenta_destino": cuenta_destino,
            "intentos": 0
        }
        _add_to_pending_queue(entry)
        logger.warning(f"⏳ Excel bloqueado, agregado a cola de reintentos ({len(_pending_entries)} pendientes)")
        return {
            "success": True,  # Reportamos éxito porque se reintentará
            "message": "Excel bloqueado - guardado en cola para reintentar automáticamente",
            "pending": True,
            "pending_count": len(_pending_entries)
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


def _add_to_pending_queue(entry: Dict[str, Any]):
    """Agrega una entrada a la cola de pendientes y arranca el thread de retry."""
    global _retry_thread, _retry_running
    
    with _pending_lock:
        _pending_entries.append(entry)
        
    # Iniciar thread de retry si no está corriendo
    if not _retry_running:
        _retry_running = True
        _retry_thread = threading.Thread(target=_retry_pending_entries, daemon=True)
        _retry_thread.start()


def _retry_pending_entries():
    """Thread que reintenta guardar las entradas pendientes cada 10 segundos."""
    global _retry_running
    
    while True:
        time.sleep(10)  # Esperar 10 segundos entre intentos
        
        with _pending_lock:
            if not _pending_entries:
                _retry_running = False
                logger.info("✅ Cola de Excel vacía, deteniendo reintentos")
                return
            
            # Tomar la primera entrada pendiente
            entry = _pending_entries[0]
        
        # Intentar guardar
        entry["intentos"] += 1
        logger.info(f"🔄 Reintentando guardar en Excel (intento #{entry['intentos']}, {len(_pending_entries)} pendientes)...")
        
        result = guardar_en_excel(
            datos=entry["datos"],
            ruta_excel=entry["ruta_excel"],
            whatsapp_from=entry["whatsapp_from"],
            timestamp_recepcion=entry["timestamp_recepcion"],
            cuenta_destino=entry["cuenta_destino"]
        )
        
        if result.get("success") and not result.get("pending"):
            # Guardado exitoso, remover de la cola
            with _pending_lock:
                if _pending_entries and _pending_entries[0] is entry:
                    _pending_entries.pop(0)
            logger.info(f"✅ Entrada guardada exitosamente en Excel (quedaban {len(_pending_entries)} pendientes)")
        elif entry["intentos"] >= 30:  # Máximo 30 intentos (5 minutos)
            # Demasiados intentos, descartar
            with _pending_lock:
                if _pending_entries and _pending_entries[0] is entry:
                    _pending_entries.pop(0)
            logger.error(f"❌ Descartando entrada después de {entry['intentos']} intentos fallidos")


def get_pending_count() -> int:
    """Retorna el número de entradas pendientes en la cola."""
    with _pending_lock:
        return len(_pending_entries)
